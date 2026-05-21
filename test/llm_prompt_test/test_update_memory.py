import sys
from pathlib import Path
from typing import Any
sys.path.insert(0, str(Path(__file__).parent))

from lifeprism.llm.prompts import prompt_loader, PromptRef, Prompts
from lifeprism.llm.providers import create_llm_client, LLMResponse
from lifeprism.llm.utils.md_os import read_md
from lifeprism.llm.agent.tools.lifeprismsystem import query_user_activity_summary
from llm_test_base import LLMTestBase, TestLog
import asyncio
import openpyxl
from datetime import datetime, timedelta

# prompts 目录路径
PROMPTS_DIR = Path(__file__).parent.parent.parent / "templates" / "prompts"


class UpdateMemoryTest(LLMTestBase):
    """记忆更新测试类"""

    def __init__(
        self,
        prompt_version: str = "v1",
        input_path: Path = Path("D:/desktop/软件开发/LifeWatch-AI/test/llm_prompt_test/dataset"),
        output_path: Path = Path("D:/desktop/软件开发/LifeWatch-AI/test/llm_prompt_test/results"),
        temperature: float = 0.7,
        start_date: str = "2026-05-13",
        end_date: str = "2026-05-19",
        prompt_params: dict[str, str] | None = None
    ):
        """
        Args:
            prompt_version: prompt 版本
            input_path: 输入数据路径
            output_path: 输出结果路径
            temperature: LLM 温度参数
            start_date: 开始日期 (YYYY-MM-DD)
            end_date: 结束日期 (YYYY-MM-DD)
            prompt_params: prompt 静态参数
        """
        super().__init__(
            prompt=Prompts.Schedule.UPDATE_MEMORY,
            prompt_version=prompt_version,
            input_path=input_path,
            output_path=output_path,
            temperature=temperature
        )
        self.llm_client = create_llm_client()
        self.prompt_params = prompt_params or {}
        self.start_date = start_date
        self.end_date = end_date

    def _get_behavior_files(self, start_date: str, end_date: str) -> list[Path]:
        """
        获取指定日期范围内的行为数据文件

        Args:
            start_date: 开始日期 (YYYY-MM-DD)
            end_date: 结束日期 (YYYY-MM-DD)

        Returns:
            行为数据文件路径列表
        """
        behavior_dir = self.input_path / "update_memory"
        if not behavior_dir.exists():
            raise FileNotFoundError(f"行为数据目录不存在: {behavior_dir}")

        start = datetime.strptime(start_date, "%Y-%m-%d")
        end = datetime.strptime(end_date, "%Y-%m-%d")

        files = []
        current = start
        while current <= end:
            date_str = current.strftime("%Y-%m-%d")
            file_path = behavior_dir / f"{date_str}.md"
            if file_path.exists():
                files.append(file_path)
            current += timedelta(days=1)

        return sorted(files)

    def _build_behavior_content(self, behavior_files: list[Path]) -> str:
        """
        构建 behavior.md 内容

        Args:
            behavior_files: 行为数据文件列表

        Returns:
            合并后的 behavior 内容
        """
        content_parts = []
        for file_path in behavior_files:
            file_content = read_md(file_path)
            if file_content:
                content_parts.append(file_content)

        return "\n\n".join(content_parts)

    def _get_computer_usage_stats(self) -> str:
        """
        获取近7天的电脑使用统计数据

        Returns:
            格式化的电脑使用统计数据
        """
        end_dt = datetime.strptime(self.end_date, "%Y-%m-%d")
        start_dt = datetime.strptime(self.start_date, "%Y-%m-%d")

        start_time = start_dt.strftime("%Y-%m-%d 00:00:00")
        end_time = end_dt.strftime("%Y-%m-%d 23:59:59")

        try:
            stats = query_user_activity_summary(
                query_option={"computer_overview"},
                start_time=start_time,
                end_time=end_time
            )
            return stats
        except Exception as e:
            print(f"获取电脑使用统计失败: {e}")
            return "## 电脑总体使用统计\n获取数据失败"

    def _build_messages(self, behavior_content: str, recent_state_content: str = "") -> list[dict[str, str]]:
        """构建 LLM 消息，使用 PromptLoader 加载 prompt"""
        # 合并静态参数和动态参数
        params = self.prompt_params.copy()
        params["recent_state_path"] = "test/output/recent_state.md"  # 测试用路径

        # 使用 PromptLoader 加载 prompt
        task_prompt = prompt_loader.load_prompt(
            self.prompt,
            version=self.prompt_version,
            **params
        )

        system_prompt = task_prompt

        # 构建用户消息
        user_prompt = f"""当前没有提供文件编辑工具，你直接输出最新的文档内容即可
        ## behavior.md
{behavior_content}

## 旧版本 recent_state.md
{recent_state_content if recent_state_content else "暂无旧版本内容"}
"""

        # v2版本需要添加电脑使用统计数据
        if self.prompt_version != "v1":
            computer_usage_stats = self._get_computer_usage_stats()
            user_prompt += f"""

## 电脑使用统计 (computer_usage_stats)
{computer_usage_stats}
"""

        return [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt}
        ]

    def _filter_behavior_by_date(self, content: str, start_date: str, end_date: str) -> str:
        """
        按日期范围筛选 behavior.md 内容

        Args:
            content: behavior.md 完整内容
            start_date: 开始日期 (YYYY-MM-DD)
            end_date: 结束日期 (YYYY-MM-DD)

        Returns:
            筛选后的 behavior 内容
        """
        import re
        
        start = datetime.strptime(start_date, "%Y-%m-%d")
        end = datetime.strptime(end_date, "%Y-%m-%d")
        
        # 按日期标题分割: ## YYYY-MM-DD
        date_pattern = re.compile(r'^## (\d{4}-\d{2}-\d{2})', re.MULTILINE)
        
        # 找到所有日期标题的位置
        matches = list(date_pattern.finditer(content))
        
        if not matches:
            return content
        
        filtered_parts = []
        
        for i, match in enumerate(matches):
            date_str = match.group(1)
            date = datetime.strptime(date_str, "%Y-%m-%d")
            
            # 检查日期是否在范围内
            if start <= date <= end:
                # 获取当前日期段的起始位置
                start_pos = match.start()
                # 获取下一个日期段的起始位置（或文件末尾）
                end_pos = matches[i + 1].start() if i + 1 < len(matches) else len(content)
                
                # 提取该日期段内容
                section = content[start_pos:end_pos].rstrip()
                filtered_parts.append(section)
        
        if not filtered_parts:
            raise ValueError(f"behavior.md 中未找到 {start_date} 到 {end_date} 范围内的数据")
        
        return "\n\n".join(filtered_parts)

    def data_input(self, input_files: list[str] | None = None) -> list[dict[str, Any]]:
        """
        解析输入数据

        Args:
            input_files: 输入的文件名称列表，None 为使用默认日期范围

        Returns:
            包含日期范围、behavior 内容和 recent_state 内容的数据列表
        """
        # 读取预生成的 behavior.md
        behavior_path = self.input_path / "update_memory" / "behavior.md"
        if not behavior_path.exists():
            raise FileNotFoundError(f"behavior.md 不存在: {behavior_path}")

        full_content = read_md(behavior_path)
        if not full_content:
            raise ValueError(f"behavior.md 内容为空: {behavior_path}")

        # 按日期范围筛选内容
        behavior_content = self._filter_behavior_by_date(full_content, self.start_date, self.end_date)

        # 读取旧版本 recent_state.md（如果存在）
        recent_state_path = self.input_path / "update_memory" / "recent_state.md"
        recent_state_content = ""
        if recent_state_path.exists():
            recent_state_content = read_md(recent_state_path)

        return [{
            "date_range": f"{self.start_date} ~ {self.end_date}",
            "behavior_content": behavior_content,
            "recent_state_content": recent_state_content,
            "file_count": 1
        }]

    async def _call_llm(self, messages: list[dict[str, str]]) -> str:
        """调用 LLM 获取响应"""
        response: LLMResponse = await self.llm_client.chat_with_retry(
            messages=messages,
            temperature=self.temperature
        )
        return response.content or ""

    async def _process_single(self, data: dict) -> tuple[dict, TestLog]:
        """处理单个数据项"""
        messages = self._build_messages(
            data["behavior_content"],
            data["recent_state_content"]
        )

        # llm_input 包含日期范围和文件数量
        llm_input = f"日期范围: {data['date_range']}\n文件数量: {data['file_count']}\n---\n行为数据长度: {len(data['behavior_content'])} 字符"
        llm_output = await self._call_llm(messages)

        test_log: TestLog = {
            "system_prompt": messages[0]["content"],
            "user_message": messages[1]["content"],
            "result": llm_output,
            "version": self.prompt_version,
            "temperature": self.temperature,
            "input_data_date": data["date_range"]
        }

        return {
            "llm_input": llm_input,
            "llm_output": llm_output,
            "date_range": data["date_range"]
        }, test_log

    async def _run_test_async(self, input_files: list[str] | None = None, round: int = 1) -> tuple[list[dict], list[TestLog]]:
        """异步执行测试"""
        data_list = self.data_input(input_files)
        results = []
        test_logs = []

        for data in data_list:
            result, log = await self._process_single(data)
            results.append(result)
            test_logs.append(log)

        return results, test_logs

    def run_test(self, input_files: list[str] | None = None, round: int = 1) -> tuple[list[dict], list[TestLog]]:
        """
        执行测试

        Args:
            input_files: 输入的文件名称列表，None 为使用默认日期范围
            round: 测试轮次

        Returns:
            测试结果列表和测试日志列表
        """
        return asyncio.run(self._run_test_async(input_files, round))

    def generate_eval_sheet(self, test_results: list[dict], round: int, temperature: float) -> Path:
        """
        生成 Excel 评估表

        Args:
            test_results: 测试结果列表
            round: 测试轮次
            temperature: 温度参数

        Returns:
            Excel 文件路径
        """
        # 创建输出目录
        version_dir = self.output_path / self.prompt.name / self.prompt_version
        version_dir.mkdir(parents=True, exist_ok=True)

        # 文件名格式: r{round}-t{temperature}.xlsx
        sheet_name = f"r{round}-t{temperature}"
        file_path = version_dir / f"{sheet_name}.xlsx"

        # 创建工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        # 写入表头
        for col_idx, col_name in enumerate(self.EVAL_SHEET_COLUMNS, 1):
            ws.cell(row=1, column=col_idx, value=col_name)

        # 写入数据
        for row_idx, result in enumerate(test_results, 2):
            ws.cell(row=row_idx, column=1, value=result.get("llm_input", ""))
            ws.cell(row=row_idx, column=2, value=result.get("llm_output", ""))
            # pass, score, reason, other 留空供人工填写

        # 调整列宽
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 80
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 20

        wb.save(file_path)
        return file_path

    def read_eval_result(self, eval_sheet_path: Path) -> float:
        """
        读取评估结果，计算通过率

        Args:
            eval_sheet_path: Excel 评估表路径

        Returns:
            通过率（0-1）
        """
        if not eval_sheet_path.exists():
            raise FileNotFoundError(f"评估表文件不存在: {eval_sheet_path}")

        wb = openpyxl.load_workbook(eval_sheet_path)
        ws = wb.active

        # 统计通过数
        total = 0
        passed = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:  # 跳过空行
                continue
            total += 1
            # "pass" 列是第3列（索引2）
            if len(row) > 2 and row[2] is not None:
                pass_value = str(row[2]).strip().lower()
                if pass_value in ("true", "1", "yes", "通过", "√"):
                    passed += 1

        if total == 0:
            return 0.0

        return passed / total

    def main(self, input_files: list[str] | None = None):
        """
        执行测试主函数

        Args:
            input_files: 输入文件列表，None 为使用默认日期范围
        """
        # 自动获取下一个轮次
        round = self.get_next_round()

        print(f"Prompt: {self.prompt}")
        print(f"Prompt 版本: {self.prompt_version}")
        print(f"Input path: {self.input_path}")
        print(f"Output path: {self.output_path}")
        print(f"Temperature: {self.temperature}")
        print(f"日期范围: {self.start_date} ~ {self.end_date}")
        print(f"Round: {round}")
        print("-" * 50)

        # 1. 执行测试
        print("开始执行测试...")
        test_results, test_logs = self.run_test(input_files=input_files, round=round)
        print(f"测试完成，共 {len(test_results)} 条结果")
        print("-" * 50)

        # 2. 生成 Excel 评估表
        print("生成 Excel 评估表...")
        eval_sheet_path = self.generate_eval_sheet(
            test_results=test_results,
            round=round,
            temperature=self.temperature
        )
        print(f"评估表已生成: {eval_sheet_path}")
        print("-" * 50)

        # 3. 保存测试日志
        print("保存测试日志...")
        self.save_log(test_logs, round)
        print(f"测试日志已保存: {self.output_path / self.prompt.name / self.prompt_version / f'r{round}-t{self.temperature}.json'}")
        print("-" * 50)

        # 4. 获取输入文件信息
        input_file_info = [f"{self.start_date}~{self.end_date}"]

        # 5. 更新 metadata（pass_ratio 初始为 0，等待人工评估）
        print("更新 metadata...")
        self.update_metadata(
            round=round,
            pass_ratio=0.0,
            input_files=input_file_info
        )
        print(f"Metadata 已更新: {self.output_path / self.prompt.name / 'meta_data.json'}")
        print("-" * 50)

        print("测试流程完成！")
        print(f"请打开评估表进行人工评估: {eval_sheet_path}")
        print("评估完成后，使用 read_eval_result() 读取通过率")

        return eval_sheet_path


if __name__ == "__main__":
    # v1 版本测试
    test = UpdateMemoryTest(
        prompt_version="v2",
        temperature=0.7,
        start_date="2026-05-13",
        end_date="2026-05-19",
        prompt_params = {"upper_limit" : 2000}
    )
    test.main()
