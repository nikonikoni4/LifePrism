import sys
from pathlib import Path
from typing import Any

sys.path.insert(0, str(Path(__file__).parent))

import asyncio

import openpyxl
from llm_test_base import LLMTestBase, TestLog

from lifeprism.llm.prompts import PromptRef, Prompts, prompt_loader
from lifeprism.llm.providers import LLMResponse, create_llm_client
from lifeprism.llm.utils.md_os import read_md

# prompts 目录路径
PROMPTS_DIR = Path(__file__).parent.parent.parent / "templates" / "prompts"


class DiarySummarySummary(LLMTestBase):
    """日记总结测试类"""

    def __init__(
        self,
        prompt_version: str = "v1",
        input_path: Path = Path("D:/desktop/软件开发/LifeWatch-AI/test/llm_prompt_test/dataset"),
        output_path: Path = Path("D:/desktop/软件开发/LifeWatch-AI/test/llm_prompt_test/results"),
        temperature: float = 0.7,
        prompt_params: dict[str, str] | None = None,
    ):
        """
        Args:
            prompt_version: prompt 版本
            input_path: 输入数据路径
            output_path: 输出结果路径
            temperature: LLM 温度参数
            prompt_params: prompt 静态参数（如 user_md），动态参数（如 upper_limit）会在运行时计算
        """
        super().__init__(
            prompt=Prompts.Schedule.CREATE_DIARY_SUMMARY,
            prompt_version=prompt_version,
            input_path=input_path,
            output_path=output_path,
            temperature=temperature,
        )
        self.llm_client = create_llm_client()
        self.prompt_params = prompt_params or {}

    def _get_diary_files(self, input_files: list[str] | None = None) -> list[Path]:
        """获取日记文件列表"""
        diary_dir = self.input_path / "diary"
        if not diary_dir.exists():
            raise FileNotFoundError(f"日记目录不存在: {diary_dir}")

        if input_files:
            return [diary_dir / f for f in input_files if (diary_dir / f).exists()]
        else:
            return sorted(diary_dir.glob("*.md"))

    def _build_messages(self, diary_content: str, date: str) -> list[dict[str, str]]:
        """构建 LLM 消息，使用 PromptLoader 加载 prompt"""
        # 合并静态参数和动态参数
        params = self.prompt_params.copy()

        if self.prompt_version != "v1":
            upper_limit = int(min(max(len(diary_content) * 0.3, 100), 500))
            params["upper_limit"] = str(upper_limit)

        # 使用 PromptLoader 加载 prompt
        task_prompt = prompt_loader.load_prompt(self.prompt, version=self.prompt_version, **params)

        system_prompt = task_prompt
        user_prompt = f"""## 日记内容：需要总结的部分
<diary>
## {date}
{diary_content}
</diary>"""

        return [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ]

    def data_input(self, input_files: list[str] | None = None) -> list[dict[str, Any]]:
        """
        解析输入数据

        Args:
            input_files: 输入的文件名称列表，None 为全量测试

        Returns:
            包含日期、内容和文件名的数据列表
        """
        diary_files = self._get_diary_files(input_files)
        data_list = []

        for file_path in diary_files:
            content = read_md(file_path)
            if content:
                date = file_path.stem  # 文件名作为日期
                data_list.append({"date": date, "content": content, "file_name": file_path.name})

        return data_list

    async def _call_llm(self, messages: list[dict[str, str]]) -> str:
        """调用 LLM 获取响应"""
        response: LLMResponse = await self.llm_client.chat_with_retry(
            messages=messages, temperature=self.temperature
        )
        return response.content or ""

    async def _process_single(self, data: dict) -> tuple[dict, TestLog]:
        """处理单个数据项"""
        messages = self._build_messages(data["content"], data["date"])
        # llm_input 包含文件标识和原始输入数据
        llm_input = f"文件: {data['file_name']}\n---\n{data['content']}"
        llm_output = await self._call_llm(messages)
        test_log: TestLog = {
            "system_prompt": messages[0]["content"],
            "user_message": messages[1]["content"],
            "result": llm_output,
            "version": self.prompt_version,
            "temperature": self.temperature,
            "input_data_date": data["date"],
        }
        return {
            "llm_input": llm_input,
            "llm_output": llm_output,
            "file_name": data["file_name"],
        }, test_log

    async def _run_test_async(
        self, input_files: list[str] | None = None, round: int = 1
    ) -> tuple[list[dict], list[TestLog]]:
        """异步执行测试，分组调用每组30个"""
        data_list = self.data_input(input_files)
        results = []
        test_logs = []
        batch_size = 30

        for i in range(0, len(data_list), batch_size):
            batch = data_list[i : i + batch_size]
            batch_results = await asyncio.gather(*[self._process_single(data) for data in batch])
            for result, log in batch_results:
                results.append(result)
                test_logs.append(log)

        return results, test_logs

    def run_test(
        self, input_files: list[str] | None = None, round: int = 1
    ) -> tuple[list[dict], list[TestLog]]:
        """
        执行测试

        Args:
            input_files: 输入的文件名称列表，None 为全量测试
            round: 测试轮次

        Returns:
            测试结果列表和测试日志列表
        """
        import asyncio

        return asyncio.run(self._run_test_async(input_files, round))

    def generate_eval_sheet(self, test_results: list[dict], round: int, temperature: float) -> Path:
        """
        生成 Excel 评估表

        Args:
            test_results: 测试结果列表
            round: 测试轮次
            temperature: 温度参数

        Returns:
            Excel 文件路径
        """
        # 创建输出目录
        version_dir = self.output_path / self.prompt.name / self.prompt_version
        version_dir.mkdir(parents=True, exist_ok=True)

        # 文件名格式: r{round}-t{temperature}.xlsx
        sheet_name = f"r{round}-t{temperature}"
        file_path = version_dir / f"{sheet_name}.xlsx"

        # 创建工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        # 写入表头
        for col_idx, col_name in enumerate(self.EVAL_SHEET_COLUMNS, 1):
            ws.cell(row=1, column=col_idx, value=col_name)

        # 写入数据
        for row_idx, result in enumerate(test_results, 2):
            ws.cell(row=row_idx, column=1, value=result.get("llm_input", ""))
            ws.cell(row=row_idx, column=2, value=result.get("llm_output", ""))
            # pass, score, reason, other 留空供人工填写

        # 调整列宽
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 60
        ws.column_dimensions["C"].width = 10
        ws.column_dimensions["D"].width = 10
        ws.column_dimensions["E"].width = 30
        ws.column_dimensions["F"].width = 20

        wb.save(file_path)
        return file_path

    def read_eval_result(self, eval_sheet_path: Path) -> float:
        """
        读取评估结果，计算通过率

        Args:
            eval_sheet_path: Excel 评估表路径

        Returns:
            通过率（0-1）
        """
        if not eval_sheet_path.exists():
            raise FileNotFoundError(f"评估表文件不存在: {eval_sheet_path}")

        wb = openpyxl.load_workbook(eval_sheet_path)
        ws = wb.active

        # 统计通过数
        total = 0
        passed = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:  # 跳过空行
                continue
            total += 1
            # "pass" 列是第3列（索引2）
            if len(row) > 2 and row[2] is not None:
                pass_value = str(row[2]).strip().lower()
                if pass_value in ("true", "1", "yes", "通过", "√"):
                    passed += 1

        if total == 0:
            return 0.0

        return passed / total

    def main(self, input_files: list[str] | None = None):
        """
        执行测试主函数

        Args:
            input_files: 输入文件列表，None 为全量测试
        """
        # 自动获取下一个轮次
        round = self.get_next_round()

        print(f"Prompt: {self.prompt}")
        print(f"Prompt 版本: {self.prompt_version}")
        print(f"Input path: {self.input_path}")
        print(f"Output path: {self.output_path}")
        print(f"Temperature: {self.temperature}")
        print(f"Round: {round}")
        print("-" * 50)

        # 1. 执行测试
        print("开始执行测试...")
        test_results, test_logs = self.run_test(input_files=input_files, round=round)
        print(f"测试完成，共 {len(test_results)} 条结果")
        print("-" * 50)

        # 2. 生成 Excel 评估表
        print("生成 Excel 评估表...")
        eval_sheet_path = self.generate_eval_sheet(
            test_results=test_results, round=round, temperature=self.temperature
        )
        print(f"评估表已生成: {eval_sheet_path}")
        print("-" * 50)

        # 3. 保存测试日志
        print("保存测试日志...")
        self.save_log(test_logs, round)
        print(
            f"测试日志已保存: {self.output_path / self.prompt.name / self.prompt_version / f'r{round}-t{self.temperature}.json'}"
        )
        print("-" * 50)

        # 4. 获取输入文件列表
        input_file_names = [r["file_name"] for r in test_results]

        # 5. 更新 metadata（pass_ratio 初始为 0，等待人工评估）
        print("更新 metadata...")
        self.update_metadata(round=round, pass_ratio=0.0, input_files=input_file_names)
        print(f"Metadata 已更新: {self.output_path / self.prompt.name / 'meta_data.json'}")
        print("-" * 50)

        print("测试流程完成！")
        print(f"请打开评估表进行人工评估: {eval_sheet_path}")
        print("评估完成后，使用 read_eval_result() 读取通过率")

        return eval_sheet_path


if __name__ == "__main__":
    # v1 版本测试（无需参数）
    # test = DiarySummarySummary(prompt_version="v1", temperature=0.7)

    # v2 版本测试（需要 user_md 参数）
    user_md = ""
    test = DiarySummarySummary(
        prompt_version="v4", temperature=0.7, prompt_params={"user_md": user_md}
    )
    test.main()
