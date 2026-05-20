import sys
from pathlib import Path
from typing import Any
import json
sys.path.insert(0, str(Path(__file__).parent))

from lifeprism.llm.prompts import prompt_loader, PromptRef, Prompts
from lifeprism.llm.providers import create_llm_client, LLMResponse
from llm_test_base import LLMTestBase, TestLog
import asyncio
import openpyxl


class ChatHistoryExtract(LLMTestBase):
    """聊天记录提取测试类"""

    def __init__(
        self,
        prompt_version: str = "v1",
        input_path: Path = Path("D:/desktop/软件开发/LifeWatch-AI/test/llm_prompt_test/dataset"),
        output_path: Path = Path("D:/desktop/软件开发/LifeWatch-AI/test/llm_prompt_test/results"),
        temperature: float = 0.7
    ):
        super().__init__(
            prompt=Prompts.Schedule.EXTRACT_CHAT,
            prompt_version=prompt_version,
            input_path=input_path,
            output_path=output_path,
            temperature=temperature
        )
        self.llm_client = create_llm_client()

    def _get_session_files(self, input_files: list[str] | None = None) -> list[Path]:
        """获取聊天会话文件列表"""
        session_dir = self.input_path / "chat_session"
        if not session_dir.exists():
            raise FileNotFoundError(f"聊天会话目录不存在: {session_dir}")

        if input_files:
            return [session_dir / f for f in input_files if (session_dir / f).exists()]
        else:
            return sorted(session_dir.glob("*.jsonl"))

    def _parse_session_file(self, file_path: Path) -> dict:
        """解析单个会话文件"""
        messages = []
        metadata = {}

        with open(file_path, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                data = json.loads(line)

                # 提取 metadata
                if data.get("_type") == "metadata":
                    metadata = {
                        "name": data.get("name", ""),
                        "created_at": data.get("created_at", ""),
                        "updated_at": data.get("updated_at", "")
                    }
                    continue

                # 提取消息内容（只处理 user 和 assistant 的文本内容）
                role = data.get("role", "")
                content = data.get("content", "")

                if role in ("user", "assistant") and content:
                    messages.append({
                        "role": role,
                        "content": content
                    })

        return {
            "metadata": metadata,
            "messages": messages
        }

    def _format_messages_for_llm(self, session_data: dict) -> str:
        """将消息格式化为 LLM 输入"""
        lines = []
        for msg in session_data["messages"]:
            role = "用户" if msg["role"] == "user" else "AI助手"
            lines.append(f"[{role}]: {msg['content']}")
        return "\n\n".join(lines)

    def _build_messages(self, chat_content: str, session_name: str) -> list[dict[str, str]]:
        """构建 LLM 消息"""
        task_prompt = prompt_loader.load_prompt(
            self.prompt,
            version=self.prompt_version
        )

        system_prompt = task_prompt
        user_prompt = f"""## 聊天记录：需要提取信息的部分
<chat_session>
会话名称: {session_name}
{chat_content}
</chat_session>"""

        return [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt}
        ]

    def data_input(self, input_files: list[str] | None = None) -> list[dict[str, Any]]:
        """
        解析输入数据

        Args:
            input_files: 输入的文件名称列表，None 为全量测试

        Returns:
            包含会话数据、格式化内容和文件名的数据列表
        """
        session_files = self._get_session_files(input_files)
        data_list = []

        for file_path in session_files:
            session_data = self._parse_session_file(file_path)
            if session_data["messages"]:
                formatted_content = self._format_messages_for_llm(session_data)
                # 从 metadata 的 created_at 提取日期 (YYYY-MM-DD)
                created_at = session_data["metadata"].get("created_at", "")
                created_at_date = created_at.split("T")[0] if created_at else ""
                data_list.append({
                    "session_name": session_data["metadata"].get("name", file_path.stem),
                    "content": formatted_content,
                    "message_count": len(session_data["messages"]),
                    "file_name": file_path.name,
                    "created_at_date": created_at_date
                })

        return data_list

    async def _call_llm(self, messages: list[dict[str, str]]) -> str:
        """调用 LLM 获取响应"""
        response: LLMResponse = await self.llm_client.chat_with_retry(
            messages=messages,
            temperature=self.temperature
        )
        return response.content or ""

    async def _process_single(self, data: dict) -> tuple[dict, TestLog]:
        """处理单个数据项"""
        messages = self._build_messages(data["content"], data["session_name"])
        # llm_input 包含文件标识和原始输入数据
        llm_input = f"文件: {data['file_name']}\n---\n{data['content']}"
        llm_output = await self._call_llm(messages)
        test_log: TestLog = {
            "system_prompt": messages[0]["content"],
            "user_message": messages[1]["content"],
            "result": llm_output,
            "version": self.prompt_version,
            "temperature": self.temperature,
            "input_data_date": data["created_at_date"]
        }
        return {
            "llm_input": llm_input,
            "llm_output": llm_output,
            "file_name": data["file_name"]
        }, test_log

    async def _run_test_async(self, input_files: list[str] | None = None, round: int = 1) -> tuple[list[dict], list[TestLog]]:
        """异步执行测试，分组调用每组30个"""
        data_list = self.data_input(input_files)
        results = []
        test_logs = []
        batch_size = 30

        for i in range(0, len(data_list), batch_size):
            batch = data_list[i:i + batch_size]
            batch_results = await asyncio.gather(
                *[self._process_single(data) for data in batch]
            )
            for result, log in batch_results:
                results.append(result)
                test_logs.append(log)

        return results, test_logs

    def run_test(self, input_files: list[str] | None = None, round: int = 1) -> tuple[list[dict], list[TestLog]]:
        """
        执行测试

        Args:
            input_files: 输入的文件名称列表，None 为全量测试
            round: 测试轮次

        Returns:
            测试结果列表和测试日志列表
        """
        return asyncio.run(self._run_test_async(input_files, round))

    def generate_eval_sheet(self, test_results: list[dict], round: int, temperature: float) -> Path:
        """
        生成 Excel 评估表

        Args:
            test_results: 测试结果列表
            round: 测试轮次
            temperature: 温度参数

        Returns:
            Excel 文件路径
        """
        version_dir = self.output_path / self.prompt.name / self.prompt_version
        version_dir.mkdir(parents=True, exist_ok=True)

        sheet_name = f"r{round}-t{temperature}"
        file_path = version_dir / f"{sheet_name}.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        for col_idx, col_name in enumerate(self.EVAL_SHEET_COLUMNS, 1):
            ws.cell(row=1, column=col_idx, value=col_name)

        for row_idx, result in enumerate(test_results, 2):
            ws.cell(row=row_idx, column=1, value=result.get("llm_input", ""))
            ws.cell(row=row_idx, column=2, value=result.get("llm_output", ""))

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 20

        wb.save(file_path)
        return file_path

    def read_eval_result(self, eval_sheet_path: Path) -> float:
        """
        读取评估结果，计算通过率

        Args:
            eval_sheet_path: Excel 评估表路径

        Returns:
            通过率（0-1）
        """
        if not eval_sheet_path.exists():
            raise FileNotFoundError(f"评估表文件不存在: {eval_sheet_path}")

        wb = openpyxl.load_workbook(eval_sheet_path)
        ws = wb.active

        total = 0
        passed = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            total += 1
            if len(row) > 2 and row[2] is not None:
                pass_value = str(row[2]).strip().lower()
                if pass_value in ("true", "1", "yes", "通过", "√"):
                    passed += 1

        if total == 0:
            return 0.0

        return passed / total

    def main(self, input_files: list[str] | None = None):
        """
        执行测试主函数

        Args:
            input_files: 输入文件列表，None 为全量测试
        """
        # 自动获取下一个轮次
        round = self.get_next_round()
        
        print(f"Prompt: {self.prompt}")
        print(f"Prompt 版本: {self.prompt_version}")
        print(f"Input path: {self.input_path}")
        print(f"Output path: {self.output_path}")
        print(f"Temperature: {self.temperature}")
        print(f"Round: {round}")
        print("-" * 50)

        print("开始执行测试...")
        test_results, test_logs = self.run_test(input_files=input_files, round=round)
        print(f"测试完成，共 {len(test_results)} 条结果")
        print("-" * 50)

        print("生成 Excel 评估表...")
        eval_sheet_path = self.generate_eval_sheet(
            test_results=test_results,
            round=round,
            temperature=self.temperature
        )
        print(f"评估表已生成: {eval_sheet_path}")
        print("-" * 50)

        print("保存测试日志...")
        self.save_log(test_logs, round)
        print(f"测试日志已保存: {self.output_path / self.prompt.name / self.prompt_version / f'r{round}-t{self.temperature}.json'}")
        print("-" * 50)

        input_file_names = [r["file_name"] for r in test_results]

        print("更新 metadata...")
        self.update_metadata(
            round=round,
            pass_ratio=0.0,
            input_files=input_file_names
        )
        print(f"Metadata 已更新: {self.output_path / self.prompt.name / 'meta_data.json'}")
        print("-" * 50)

        print("测试流程完成！")
        print(f"请打开评估表进行人工评估: {eval_sheet_path}")
        print("评估完成后，使用 read_eval_result() 读取通过率")

        return eval_sheet_path


if __name__ == "__main__":
    # 默认值
    default_version = "v1"
    default_temperature = 0.7
    
    # 交互式输入
    version_input = input(f"请输入 Prompt 版本 (默认 {default_version}): ").strip()
    version = version_input if version_input else default_version
    
    temp_input = input(f"请输入 Temperature (默认 {default_temperature}): ").strip()
    temperature = float(temp_input) if temp_input else default_temperature
    
    print(f"\n使用配置: version={version}, temperature={temperature}")
    print("=" * 50)
    
    test = ChatHistoryExtract(prompt_version=version, temperature=temperature)
    test.main()
